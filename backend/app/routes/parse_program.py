from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from typing import Optional
import io
import json
import openpyxl
from google import genai

from ..auth import get_current_coach
from ..models import User
from ..config import settings

router = APIRouter(prefix="/api/coaches", tags=["coaches"])

MAX_TEXT_CHARS = 50_000

SYSTEM_INSTRUCTION = """
You are an expert strength and conditioning coach assistant with deep
knowledge of how training programs are structured and documented.

You will be given raw spreadsheet content from a workout program.
Your job is to parse it into structured JSON.

UNDERSTANDING PROGRAM STRUCTURE:
Strength programs are always hierarchical. Before parsing anything,
read the entire spreadsheet and identify the hierarchy. It may have
2, 3, or 4 levels depending on the program — infer from context:

- Top level: Training phases or blocks (often separate tabs/sheets,
  or major named sections). Each tab is almost always a separate block.
- Mid level: Time periods within a block — often weeks, but could be
  sessions, phases, or cycles. These are frequently laid out as
  REPEATING COLUMN GROUPS side by side (e.g. the same exercise row
  appears once, but columns repeat for Week 1 | Week 2 | Week 3 with
  different sets/reps/loads each time).
- Day level: Individual training sessions within a time period —
  labeled Day 1/2/3, Monday/Wednesday/Friday, Session A/B, etc.
- Block level: Named groupings of exercises within a single day —
  like Warm Up, Strength, Power, Accessory, Conditioning, Flexibility.
  These appear as row headers with no sets/reps data.

CRITICAL: When weeks are laid out as column groups, each week-day
combination is a SEPARATE workout. "Week 1 Day 1" and "Week 2 Day 1"
are different workouts with the same exercises but different
sets/reps/loads. Generate one workout entry per unique week-day pair.

WHAT IS AN EXERCISE vs. A HEADER:
- Exercise rows: have a name AND sets/reps in at least one column group
- Section headers: have a label but NO sets/reps anywhere in the row.
  Never include these as exercises.
- The block name a group of exercises belongs to (e.g. "Strength 1")
  should be included in each exercise's coach_notes as a prefix:
  "[Strength 1] original notes here"

HANDLING ANY FORMAT:
- Do not assume fixed column positions or row patterns
- Infer structure from context: indentation, spacing, empty rows,
  color hints in text, keyword patterns
- Labels may vary wildly — "Phase A", "Block 1", "GPP", "Week A",
  "Monday" all mean the same structural concept
- Process ALL tabs/sheets — each one is part of the program
- Loads may be lbs, kg, velocity (m/s), RPE, percentage, or
  descriptive text — always preserve in coach_notes exactly as written
- If reps are freeform, set reps to 1 and store full description
  in coach_notes
- Never hallucinate exercises. Only include what is in the spreadsheet.
- Rest times are especially important in rehab and PT programs.
  Always capture rest_seconds when stated. Convert to seconds:
  "30s" → 30, "1 min" or "1m" → 60, "90s" → 90,
  "No rest" or "none" → 0. If not stated, use null.
- Return ONLY valid JSON. No markdown, no backticks, no preamble.
  First character must be '{'.
"""


def _cell_to_str(val) -> str:
    if val is None:
        return ""
    # Skip formula objects that slip through with data_only=True
    if "Formula" in type(val).__name__:
        return ""
    return str(val).strip()


def _read_pdf(contents: bytes) -> str:
    import fitz
    doc = fitz.open(stream=contents, filetype="pdf")
    sections = []
    for page_num, page in enumerate(doc):
        sections.append(f"=== Page {page_num + 1} ===")
        sections.append(page.get_text())
    return "\n".join(sections)


def _read_xlsx(contents: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    sections: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sections.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows():
            row_vals = [_cell_to_str(cell.value) for cell in row]
            if any(row_vals):
                sections.append(" | ".join(row_vals))
            else:
                sections.append("")  # preserve empty rows as blank lines
        sections.append("")  # blank line between sheets
    return "\n".join(sections)


@router.post("/programs/parse")
async def parse_program(
    file: UploadFile = File(...),
    feedback_issues: Optional[str] = Form(None),
    feedback_text: Optional[str] = Form(None),
    previous_result: Optional[str] = Form(None),
    current_coach: User = Depends(get_current_coach),
):
    """Parse an Excel workout program file using Gemini AI."""
    filename = file.filename or ""
    ext = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""

    if ext not in (".xlsx", ".pdf"):
        raise HTTPException(
            status_code=422,
            detail=f"Unsupported file type '{ext or 'unknown'}'. Please upload a .xlsx or .pdf file.",
        )

    contents = await file.read()

    try:
        spreadsheet_text = _read_pdf(contents) if ext == ".pdf" else _read_xlsx(contents)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not read file: {exc}")

    if len(spreadsheet_text) > MAX_TEXT_CHARS:
        spreadsheet_text = (
            spreadsheet_text[:MAX_TEXT_CHARS]
            + "\n[... content truncated at 50,000 characters ...]"
        )

    # Build the user message
    user_message = f"""File name: {filename}

Here is the raw spreadsheet content:

{spreadsheet_text}
"""

    if feedback_issues or feedback_text:
        user_message += f"""
---
A coach reviewed the previous parse attempt and reported these issues:
Issues flagged: {feedback_issues or "None specified"}
Additional comments: {feedback_text or "None"}

Here is the previous parse result that had these problems:
{previous_result or "Not provided"}

Please re-parse the original spreadsheet carefully, specifically \
addressing every flagged issue. Do not repeat the same mistakes.
"""

    user_message += """
Return a JSON object with this exact structure:
{
  "program_name": "string — infer from filename, sheet names, or content. Use filename as fallback.",
  "description": "string — 1-2 sentence summary of the program",
  "workouts": [
    {
      "name": "string — must uniquely identify the week AND day, e.g. 'Block A - Week 1 - Day 1' or '6 Week Phase - Week 2 - Day 3'. Never just 'Day 1' alone since the same day repeats across weeks.",
      "day_offset": integer — 0-indexed sequential day number across the entire program (0, 1, 2...),
      "description": "string or null — purpose of this workout session if inferable",
      "exercises": [
        {
          "name": "string — clean exercise name only, no sets/reps or load info in the name",
          "sets": integer — must be a whole number,
          "reps": integer — use 1 if time/distance/freeform based,
          "coach_notes": "string or null — all coaching cues, load prescriptions, freeform rep descriptions, velocity targets, RPE, percentage info",
          "order": integer — 0-indexed position within this workout,
          "rest_seconds": integer or null — only if explicitly stated
        }
      ]
    }
  ]
}"""

    # Call Gemini
    try:
        client = genai.Client(api_key=settings.gemini_api_key)
        response = client.models.generate_content(
            model="gemini-2.5-flash-lite",
            contents=user_message,
            config={"system_instruction": SYSTEM_INSTRUCTION},
        )
        raw_text = response.text.strip()
    except Exception:
        raise HTTPException(
            status_code=503,
            detail="AI service unavailable — please try again",
        )

    # Strip accidental markdown fences
    if raw_text.startswith("```"):
        raw_text = raw_text.split("\n", 1)[-1]
        if raw_text.endswith("```"):
            raw_text = raw_text[:-3].rstrip()

    try:
        parsed = json.loads(raw_text)
    except json.JSONDecodeError:
        raise HTTPException(
            status_code=422,
            detail="AI returned invalid response — please try again",
        )

    return parsed
